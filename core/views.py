from django.shortcuts import render
from django.db.models import Sum
from django.core.mail import EmailMessage
from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import requests
from core.pdf_utils import generate_receipt
from .receipts import generate_receipt_pdf
from django.contrib.auth.views import LoginView
from django.contrib.auth import authenticate, login
from django.shortcuts import render, redirect
from django.contrib import messages
import os
from django.views import View
from django.template.loader import get_template  # Добавьте этот импорт
from django.utils.timezone import now
from io import BytesIO
from django.http import HttpResponse
from xhtml2pdf import pisa
from django.db.models import Sum
from django.core.files.base import ContentFile
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.urls import reverse, reverse_lazy
from .models import AcademicYear, Discount, Employee, News, SalaryPayment, Student, StudentYearContract, Grade, Income, Expense, Reservation, AuditLog, Student2, Teacher
from .forms import DiscountForm, EmployeeForm, SalaryPaymentForm, StudentForm, IncomeForm, ExpenseForm, ReservationForm
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from django.contrib.auth import logout
# views.py
from django.shortcuts import render, redirect
from .models import Application
from django.views.generic import ListView, DetailView
from .models import News
from .models import TelegramSubscriber
# Главная страница
from django.shortcuts import render
from .models import Student, AcademicYear
from django.db.models import Count, Sum, Q
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter  
from django.utils import timezone
from django.contrib import messages
from django.shortcuts import get_object_or_404
from django.views.generic import View
class NewsListView(ListView):
    model = News
    template_name = 'news_list.html'
    context_object_name = 'news_list'
    paginate_by = 6  # Показывать по 6 новостей на странице
    
    def get_queryset(self):
        return News.objects.all().order_by('-created_at')

class NewsDetailView(DetailView):
    model = News
    template_name = 'news_detail.html'
    context_object_name = 'news'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Добавляем последние новости для боковой колонки
        context['recent_news'] = News.objects.exclude(id=self.object.id).order_by('-created_at')[:3]
        return context
    
TELEGRAM_BOT_TOKEN = "7392373379:AAFmvBHQE6uCWJ817i9H3M9fKEYgUwaNoaE"

def best_student(request):
    students = Student2.objects.all()
    return render(request, 'best_student.html',{'students':students})

def application_view(request):
    if request.method == 'POST':
        child_name = request.POST.get('child_name')
        child_surname = request.POST.get('child_surname')
        child_class = request.POST.get('child_class')
        parent_phone = request.POST.get('parent_phone')
        
        Application.objects.create(
            child_name=child_name,
            child_surname=child_surname,
            child_class=child_class,
            parent_phone=parent_phone
        )
         # Формируем сообщение
        message = (
            f"📥 Новая заявка:\n"
            f"👶 Ребёнок: {child_name} {child_surname}\n"
            f"📚 Класс: {child_class}\n"
            f"📞 Телефон родителя: {parent_phone}"
        )

        # Отправка всем активным подписчикам
        subscribers = TelegramSubscriber.objects.filter(is_active=True)
        for subscriber in subscribers:
            try:
                requests.post(
                    f'https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage',
                    data={'chat_id': subscriber.chat_id, 'text': message}
                )
            except Exception as e:
                print(f"Ошибка при отправке для {subscriber.chat_id}: {e}")

    teachers = Teacher.objects.filter(is_publish=True)
     # Проверяем, есть ли главный учитель для отображения по умолчанию
    main_teacher = teachers.filter(is_main=True).first()
    if not main_teacher and teachers.exists():
        main_teacher = teachers.first()
    best_students = Student2.objects.filter(is_featured=True).order_by('order')[:5]  # Ограничив
    news_list = News.objects.filter(is_published=True).order_by('-created_at')[:3]
    return render(request, 'index.html',{'main_teacher':main_teacher, 
                                         'teachers':teachers,
                                         'best_students': best_students,
                                          'news_list': news_list,})
from django.http import JsonResponse
from django.shortcuts import get_object_or_404

def get_teacher(request, teacher_id):
    teacher = get_object_or_404(Teacher, id=teacher_id)
    
    data = {
        'success': True,
        'teacher': {
            'id': teacher.id,
            'first_name': teacher.first_name,
            'last_name': teacher.last_name,
            'subject': teacher.subject,
            'description': teacher.description,
            'image': teacher.image.url,
        }
    }
    return JsonResponse(data)

class ClassDebtsReportView(LoginRequiredMixin, View):
    def test_func(self):
        return is_admin(self.request.user) or is_accountant(self.request.user)
    
    def get(self, request):
        # Получаем текущий учебный год
        current_year = AcademicYear.objects.filter(is_current=True).first()
        if not current_year:
            return HttpResponse("Текущий учебный год не установлен", status=400)

        # Создаем Excel-файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Задолженности по классам"
        
        # Стили
        header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        header_font = Font(bold=True)
        center_aligned = Alignment(horizontal='center')
        right_aligned = Alignment(horizontal='right')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Заголовки
        headers = [
            'Класс', 'Ученик', 'Сумма контракта',
            'Скидка', 'К оплате (с учётом скидки)',
            'Оплачено', 'Остаток', 'Статус'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_aligned
            cell.border = border

        # Получаем данные по классам
        grades = Grade.objects.all().order_by('number', 'parallel')
        row_num = 2
        
        for grade in grades:
            students = grade.student_set.filter(is_active=True, is_graduated=False).filter(
                Q(enrollment_year__isnull=True) |
                Q(enrollment_year__start_date__lte=current_year.start_date)
            )
            
            for student in students:
                # Рассчитываем оплаты за текущий год
                payments = Income.objects.filter(
                    student=student,
                    academic_year=current_year,
                    status='paid'
                ).aggregate(total=Sum('amount'))['total'] or 0

                discount = Discount.objects.filter(
                    student=student,
                    academic_year=current_year,
                ).aggregate(total=Sum('amount'))['total'] or 0

                from decimal import Decimal as _D
                contract_amount = get_contract_amount(student, current_year)
                discount = _D(str(discount))
                payments = _D(str(payments))
                effective = max(contract_amount - discount, _D('0'))
                remaining = max(effective - payments, _D('0'))

                # Добавляем данные в таблицу
                ws.cell(row=row_num, column=1,
                       value=f"{grade.number}{grade.parallel}").border = border
                ws.cell(row=row_num, column=2,
                       value=student.full_name).border = border
                ws.cell(row=row_num, column=3,
                       value=contract_amount).border = border
                ws.cell(row=row_num, column=4,
                       value=discount).border = border
                ws.cell(row=row_num, column=5,
                       value=effective).border = border
                ws.cell(row=row_num, column=6,
                       value=payments).border = border
                ws.cell(row=row_num, column=7,
                       value=remaining).border = border
                ws.cell(row=row_num, column=8,
                       value=student.get_status_display()).border = border

                # Форматируем числовые ячейки
                for col in [3, 4, 5, 6, 7]:
                    ws.cell(row=row_num, column=col).number_format = '#,##0.00'
                    ws.cell(row=row_num, column=col).alignment = right_aligned

                # Подсветка должников
                if remaining > 0:
                    for col in range(1, 9):
                        ws.cell(row=row_num, column=col).fill = PatternFill(
                            start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                
                row_num += 1
        
        # Добавляем итоги по классам
        ws.cell(row=row_num, column=1, value="ИТОГО:").font = header_font
        for col, formula in [
            (3, f"SUM(C2:C{row_num-1})"),
            (4, f"SUM(D2:D{row_num-1})"),
            (5, f"SUM(E2:E{row_num-1})"),
            (6, f"SUM(F2:F{row_num-1})"),
            (7, f"SUM(G2:G{row_num-1})"),
        ]:
            ws.cell(row=row_num, column=col, value=formula)
            ws.cell(row=row_num, column=col).number_format = '#,##0.00'
            ws.cell(row=row_num, column=col).font = header_font
            ws.cell(row=row_num, column=col).alignment = right_aligned
            ws.cell(row=row_num, column=col).border = border

        # Настраиваем ширину столбцов
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

        # Формируем ответ
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename="class_debts_report_{current_year.year}.xlsx"'},
        )
        wb.save(response)
        return response
   


def logout_view(request):
    logout(request)
    return redirect('home') 

def is_admin(user):
    return user.is_superuser


def get_contract_amount(student, year):
    """Возвращает сумму контракта ученика на конкретный год.
    Если для года задана индивидуальная сумма — берёт её, иначе берёт базовую."""
    from decimal import Decimal
    if year:
        yc = StudentYearContract.objects.filter(student=student, academic_year=year).first()
        if yc:
            return yc.amount
    return student.contract_amount or Decimal('0')


@login_required
def set_current_year(request):
    if request.method == 'POST' and (request.user.is_staff or request.user.is_superuser):
        year_id = request.POST.get('year_id')
        if year_id:
            year = get_object_or_404(AcademicYear, pk=year_id)
            AcademicYear.objects.filter(is_current=True).update(is_current=False)
            year.is_current = True
            year.save()
            messages.success(request, f"Текущий учебный год изменён на {year.year}.")
    return redirect(request.POST.get('next', 'home'))

def is_accountant(user):
    return user.groups.filter(name='Бухгалтер').exists()



@login_required()
def home(request):
    from decimal import Decimal
    current_year = AcademicYear.objects.filter(is_current=True).first()

    students = Student.objects.filter(is_graduated=False)
    if current_year:
        students = students.filter(
            Q(enrollment_year__isnull=True) |
            Q(enrollment_year__start_date__lte=current_year.start_date)
        )
    studying = students.filter(status='studying')
    reserve = students.filter(status='reserve')
    expelled = students.filter(status='expelled')

    male_count = studying.filter(pol='male').count()
    female_count = studying.filter(pol='female').count()

    fully_paid = studying.filter(current_year_paid=True).count()
    not_paid = studying.filter(current_year_paid=False).count()

    total_students = studying.count()
    total_reserve = reserve.count()
    total_expelled = expelled.count()

    # Суммируем год-специфичные контракты (если есть) + базовые для остальных
    if current_year:
        from django.db.models import OuterRef, Subquery, DecimalField as DF
        from django.db.models.functions import Coalesce as C2
        yc_sub = StudentYearContract.objects.filter(
            student=OuterRef('pk'), academic_year=current_year
        ).values('amount')[:1]
        total_contract_amount = studying.annotate(
            _yc=C2(Subquery(yc_sub, output_field=DF(max_digits=10, decimal_places=2)),
                   'contract_amount', Decimal('0'))
        ).aggregate(total=Sum('_yc'))['total'] or Decimal('0')
    else:
        total_contract_amount = studying.aggregate(total=Sum('contract_amount'))['total'] or Decimal('0')
    total_paid_amount = Income.objects.filter(
        student__in=studying,
        academic_year=current_year,
        status='paid'
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
    total_discounts = Discount.objects.filter(
        student__in=studying,
        academic_year=current_year,
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
    effective_total_contract = max(total_contract_amount - total_discounts, Decimal('0'))
    total_remaining = max(effective_total_contract - total_paid_amount, Decimal('0'))

    payment_percent = 0
    if effective_total_contract > 0:
        payment_percent = int((total_paid_amount / effective_total_contract) * 100)

    # Последние 6 платежей
    recent_payments = Income.objects.filter(
        academic_year=current_year,
        status='paid'
    ).select_related('student').order_by('-date')[:6]

    # Доходы по месяцам (сентябрь–май)
    month_order = [9, 10, 11, 12, 1, 2, 3, 4, 5]
    month_names = {9:'Сен', 10:'Окт', 11:'Ноя', 12:'Дек', 1:'Янв', 2:'Фев', 3:'Мар', 4:'Апр', 5:'Май'}
    monthly_labels = []
    monthly_data = []
    if current_year:
        for m in month_order:
            if m >= 9:
                year = current_year.start_date.year
            else:
                year = current_year.end_date.year
            total = Income.objects.filter(
                academic_year=current_year,
                status='paid',
                date__year=year,
                date__month=m
            ).aggregate(total=Sum('amount'))['total'] or 0
            monthly_labels.append(month_names[m])
            monthly_data.append(float(total))

    # Скидки за текущий год
    discounts_this_year = Discount.objects.filter(
        academic_year=current_year,
        student__in=studying,
    ).select_related('student', 'student__grade').order_by('-amount')

    # Последние расходы
    recent_expenses = Expense.objects.order_by('-date')[:5]

    # Расходы за текущий год
    total_expenses = 0
    if current_year:
        total_expenses = Expense.objects.filter(
            date__gte=current_year.start_date,
            date__lte=current_year.end_date
        ).aggregate(total=Sum('amount'))['total'] or 0

    net_profit = float(total_paid_amount) - float(total_expenses)

    context = {
        'current_year': current_year,
        'total_students': total_students,
        'male_count': male_count,
        'female_count': female_count,
        'fully_paid': fully_paid,
        'not_paid': not_paid,
        'total_reserve': total_reserve,
        'total_expelled': total_expelled,
        'total_contract_amount': total_contract_amount,
        'total_discounts': total_discounts,
        'effective_total_contract': effective_total_contract,
        'total_paid_amount': total_paid_amount,
        'total_remaining': total_remaining,
        'payment_percent': payment_percent,
        'recent_payments': recent_payments,
        'monthly_labels': monthly_labels,
        'monthly_data': monthly_data,
        'recent_expenses': recent_expenses,
        'total_expenses': total_expenses,
        'net_profit': net_profit,
        'discounts_this_year': discounts_this_year,
    }

    return render(request, 'school/home.html', context)



from django.shortcuts import render
from django.views.generic import ListView
from .models import Student, Grade
# Студенты
class StudentListView(ListView):
    model = Student
    template_name = 'school/students/list.html'
    context_object_name = 'students'
    paginate_by = 20

    def get_queryset(self):
        from django.db.models import OuterRef, Subquery, DecimalField
        from django.db.models.functions import Coalesce
        from decimal import Decimal

        current_year = AcademicYear.objects.filter(is_current=True).first()

        queryset = super().get_queryset().select_related('grade')
        grade_filter = self.request.GET.get('grade')
        show_graduated = self.request.GET.get('graduated') == '1'
        search_query = self.request.GET.get('q', '').strip()

        if not show_graduated:
            queryset = queryset.filter(is_graduated=False)

        if grade_filter and '-' in grade_filter:
            number, parallel = grade_filter.split('-')
            queryset = queryset.filter(grade__number=number, grade__parallel=parallel)

        if search_query:
            queryset = queryset.filter(
                Q(full_name__icontains=search_query) |
                Q(parent_contacts__icontains=search_query)
            )

        # Аннотируем оплаты, скидки и год-специфичную сумму контракта
        from django.db.models import ExpressionWrapper, F, Value
        dec_field = DecimalField(max_digits=10, decimal_places=2)
        if current_year:
            paid_sub = Income.objects.filter(
                student=OuterRef('pk'),
                academic_year=current_year,
                status='paid'
            ).values('student').annotate(s=Sum('amount')).values('s')
            discount_sub = Discount.objects.filter(
                student=OuterRef('pk'),
                academic_year=current_year,
            ).values('student').annotate(s=Sum('amount')).values('s')
            # Год-специфичная сумма контракта (если нет — берём базовую)
            year_contract_sub = StudentYearContract.objects.filter(
                student=OuterRef('pk'),
                academic_year=current_year,
            ).values('amount')[:1]
            queryset = queryset.annotate(
                paid_this_year=Coalesce(Subquery(paid_sub, output_field=dec_field), Decimal('0')),
                discount_this_year=Coalesce(Subquery(discount_sub, output_field=dec_field), Decimal('0')),
                year_contract_amount=Coalesce(
                    Subquery(year_contract_sub, output_field=dec_field),
                    F('contract_amount'),
                    Decimal('0'),
                ),
            ).annotate(
                effective_contract=ExpressionWrapper(
                    F('year_contract_amount') - F('discount_this_year'),
                    output_field=dec_field
                )
            )
        else:
            queryset = queryset.annotate(
                paid_this_year=Value(Decimal('0'), output_field=dec_field),
                discount_this_year=Value(Decimal('0'), output_field=dec_field),
                year_contract_amount=Coalesce(F('contract_amount'), Value(Decimal('0'), output_field=dec_field)),
                effective_contract=Coalesce(F('contract_amount'), Value(Decimal('0'), output_field=dec_field)),
            )

        return queryset.order_by('grade__number', 'grade__parallel', 'full_name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['grades'] = Grade.objects.all().order_by('number', 'parallel')
        context['selected_grade'] = self.request.GET.get('grade', '')
        context['show_graduated'] = self.request.GET.get('graduated') == '1'
        context['current_year'] = AcademicYear.objects.filter(is_current=True).first()
        context['search_query'] = self.request.GET.get('q', '')
        return context


class StudentDetailView(LoginRequiredMixin, DetailView):
    model = Student
    template_name = 'school/students/detail.html'

    def get_context_data(self, **kwargs):
        from decimal import Decimal
        context = super().get_context_data(**kwargs)
        student = self.object

        current_year = AcademicYear.objects.filter(is_current=True).first()

        payments = Income.objects.filter(
            student=student,
            academic_year=current_year
        ).order_by('-date')

        total_payments = payments.filter(status='paid').aggregate(Sum('amount'))['amount__sum'] or Decimal('0')

        # Сумма контракта: год-специфичная или базовая
        contract_amount = get_contract_amount(student, current_year)

        # Скидка за текущий год
        discount_current_year = Discount.objects.filter(
            student=student, academic_year=current_year
        ).aggregate(s=Sum('amount'))['s'] or Decimal('0')
        effective_contract = max(contract_amount - discount_current_year, Decimal('0'))

        remaining_payment = max(effective_contract - total_payments, Decimal('0'))

        payment_percent = 0
        if effective_contract > 0:
            payment_percent = min(int((total_payments / effective_contract) * 100), 100)

        # Год-специфичная запись контракта для текущего года (если есть)
        year_contract = StudentYearContract.objects.filter(
            student=student, academic_year=current_year
        ).first()

        # Все скидки (сгруппированные по году для отображения в истории)
        discounts = Discount.objects.filter(student=student).select_related('academic_year').order_by('-academic_year__start_date')

        # Все платежи по всем годам (для истории)
        all_payments = Income.objects.filter(student=student).select_related('academic_year').order_by('-date')

        context.update({
            'student': student,
            'payments': payments,
            'all_payments': all_payments,
            'total_payments': total_payments,
            'contract_amount': contract_amount,
            'year_contract': year_contract,
            'discount_current_year': discount_current_year,
            'effective_contract': effective_contract,
            'remaining_payment': remaining_payment,
            'payment_percent': payment_percent,
            'current_year': current_year,
            'discounts': discounts,
            'is_fully_paid': remaining_payment <= 0 and effective_contract > 0,
        })
        return context
  
class StudentCreateView(LoginRequiredMixin, CreateView):
    model = Student
    form_class = StudentForm
    template_name = 'school/students/form.html'
    success_url = reverse_lazy('student-list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Добавление нового ученика'
        context['submit_text'] = 'Создать ученика'
        return context
    
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        # Для обычных пользователей (у них скрыто поле даты) ставим сегодняшнюю дату
        if not (self.request.user.is_staff or self.request.user.is_superuser):
            from django.utils import timezone
            form.instance.contract_date = timezone.localdate()
        response = super().form_valid(form)
        # При создании фиксируем начальную сумму как год-специфичный контракт
        contract_amount = form.cleaned_data.get('contract_amount')
        if contract_amount:
            current_year = AcademicYear.objects.filter(is_current=True).first()
            if current_year:
                StudentYearContract.objects.get_or_create(
                    student=self.object,
                    academic_year=current_year,
                    defaults={'amount': contract_amount},
                )
        messages.success(self.request, f"Ученик {self.object.full_name} успешно добавлен!")
        return response


def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            return redirect('home')  # укажи куда перенаправить после входа
        else:
            messages.error(request, 'Неверное имя пользователя или пароль.')

    return render(request, 'login.html')

class CustomLoginView(LoginView):
    template_name = 'accounts/login.html'  # путь к твоему шаблону
    redirect_authenticated_user = True  # если пользователь уже вошёл, перекинуть

    def get_success_url(self):
        return self.get_redirect_url() or '/'  # куда перекидывать после входа (например, на главную)



class StudentUpdateView(LoginRequiredMixin, UpdateView):
    model = Student
    form_class = StudentForm
    template_name = 'school/students/form.html'
    success_url = reverse_lazy('student-list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        from decimal import Decimal
        new_amount = form.cleaned_data.get('contract_amount')
        # Запоминаем текущую базовую сумму ДО сохранения
        original_amount = Student.objects.filter(pk=self.object.pk).values_list('contract_amount', flat=True).first()

        response = super().form_valid(form)  # сохраняет форму (в т.ч. contract_amount в Student)

        if new_amount is not None and self.request.user.is_superuser:
            current_year = AcademicYear.objects.filter(is_current=True).first()
            if current_year:
                # Сохраняем год-специфичный контракт для текущего года
                StudentYearContract.objects.update_or_create(
                    student=self.object,
                    academic_year=current_year,
                    defaults={'amount': new_amount},
                )
            # Восстанавливаем базовую сумму — она не должна меняться при редактировании
            Student.objects.filter(pk=self.object.pk).update(contract_amount=original_amount)

        messages.success(self.request, f"Данные ученика {self.object.full_name} обновлены.")
        return response


class StudentYearContractUpdateView(LoginRequiredMixin, View):
    """Устанавливает/обновляет сумму контракта ученика на указанный год."""

    def post(self, request, pk):
        from decimal import Decimal, InvalidOperation
        student = get_object_or_404(Student, pk=pk)
        if not request.user.is_superuser:
            messages.error(request, "Только суперпользователь может изменять сумму контракта.")
            return redirect('student-detail', pk=pk)

        year_id = request.POST.get('year_id')
        amount_raw = request.POST.get('amount', '').strip()
        try:
            amount = Decimal(amount_raw)
            if amount <= 0:
                raise ValueError
        except (InvalidOperation, ValueError):
            messages.error(request, "Укажите корректную сумму контракта.")
            return redirect('student-detail', pk=pk)

        year = get_object_or_404(AcademicYear, pk=year_id)
        StudentYearContract.objects.update_or_create(
            student=student,
            academic_year=year,
            defaults={'amount': amount},
        )
        messages.success(request, f"Сумма контракта на {year.year} изменена: {amount:,.0f} сом.")
        return redirect('student-detail', pk=pk)


class DiscountCreateView(LoginRequiredMixin, CreateView):
    model = Discount
    form_class = DiscountForm
    template_name = 'school/students/discount_form.html'

    def get_student(self):
        return get_object_or_404(Student, pk=self.kwargs['pk'])

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['student'] = self.get_student()
        return ctx

    def form_valid(self, form):
        form.instance.student = self.get_student()
        messages.success(self.request, "Скидка успешно добавлена.")
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('student-detail', kwargs={'pk': self.kwargs['pk']})


class DiscountDeleteView(LoginRequiredMixin, DeleteView):
    model = Discount
    template_name = 'school/students/discount_confirm_delete.html'

    def get_success_url(self):
        return reverse_lazy('student-detail', kwargs={'pk': self.object.student.pk})

    def form_valid(self, form):
        messages.success(self.request, "Скидка удалена.")
        return super().form_valid(form)


class PromoteGradesView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Переводит всех активных учеников на класс выше. 12-классники становятся выпускниками."""
    template_name = 'school/students/promote_grades_confirm.html'

    def test_func(self):
        return self.request.user.is_superuser

    def _build_preview(self):
        students = (
            Student.objects
            .filter(is_active=True, is_graduated=False, status='studying')
            .select_related('grade')
            .order_by('grade__number', 'grade__parallel', 'full_name')
        )
        to_graduate, to_promote, no_next_grade = [], [], []
        for student in students:
            if student.grade.number >= 12:
                to_graduate.append(student)
            else:
                next_grade = Grade.objects.filter(
                    number=student.grade.number + 1,
                    parallel=student.grade.parallel,
                ).first()
                if next_grade:
                    to_promote.append((student, next_grade))
                else:
                    no_next_grade.append(student)
        return to_graduate, to_promote, no_next_grade

    def get(self, request):
        to_graduate, to_promote, no_next_grade = self._build_preview()
        return render(request, self.template_name, {
            'to_graduate': to_graduate,
            'to_promote': to_promote,
            'no_next_grade': no_next_grade,
        })

    def post(self, request):
        to_graduate, to_promote, no_next_grade = self._build_preview()

        for student in to_graduate:
            Student.objects.filter(pk=student.pk).update(
                is_graduated=True,
                is_active=False,
            )

        for student, next_grade in to_promote:
            Student.objects.filter(pk=student.pk).update(grade=next_grade)

        messages.success(
            request,
            f"Перевод завершён: переведено {len(to_promote)}, "
            f"выпущено {len(to_graduate)}, "
            f"пропущено {len(no_next_grade)} (нет следующего класса)."
        )
        return redirect('student-list')


from django.http import HttpResponse, JsonResponse
from django.db.models import Q

def student_search(request):
    query = request.GET.get('q', '')
    if query:
        students = Student.objects.filter(
            Q(full_name__icontains=query) |
            Q(parent_contacts__icontains=query)
        ).filter(is_graduated=False)[:10]
        results = [
            {
                'id': student.id,
                'text': f"{student.full_name} ({student.grade.number}{student.grade.parallel})"
            } for student in students
        ]
    else:
        results = []
    return JsonResponse({'results': results})
# Доходы
from django.db.models import Q
# from datetime import datetime
from django.http import HttpResponse
import csv
from django.db.models import Q, Sum
# from datetime import datetime
from django.http import HttpResponse
import csv

class IncomeListView(LoginRequiredMixin, ListView):
    model = Income
    template_name = 'school/incomes/list.html'
    context_object_name = 'incomes'
    paginate_by = 20
    
    def test_func(self):
        return is_admin(self.request.user) or is_accountant(self.request.user)
    def generate_transfer_act_pdf(self, queryset):
        total_amount = queryset.aggregate(total=Sum('amount'))['total'] or 0
        date_from = self.request.GET.get('date_from', 'не указана')
        date_to = self.request.GET.get('date_to', 'не указана')
    
        context = {
        'incomes': queryset,
        'total_amount': total_amount,
        'date_from': date_from,
        'date_to': date_to,
        'generated_date': now().strftime('%d.%m.%Y %H:%M'),
        'user': self.request.user.get_full_name() or self.request.user.username,
        }
    
        template = get_template('school/incomes/transfer_act_html.html')
        html = template.render(context)
    
        return HttpResponse(html)
       

    def render_to_response(self, context, **response_kwargs):
        # Обработка экспорта в Excel
        if self.request.GET.get('export') == 'xlsx':
            return self.export_to_excel()
        # Обработка генерации акта передачи
        elif self.request.GET.get('export') == 'pdf':
            queryset = self.get_queryset()
            return self.generate_transfer_act_pdf(queryset)
            
        return super().render_to_response(context, **response_kwargs)
    def export_to_excel(self):
        queryset = self.get_queryset()
        total_amount = queryset.aggregate(Sum('amount'))['amount__sum'] or 0

        # Создаем Excel-файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Приходы"

        # Стили для заголовков
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Заголовки
        headers = [
            'Дата', 'Ученик', 'Класс', 'Сумма (сом)', 
            'Способ оплаты', 'Статус', 'Номер транзакции', 'Период оплаты'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border

        # Данные
        for row_num, income in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1, value=income.date.strftime('%d.%m.%Y')).border = border
            ws.cell(row=row_num, column=2, value=income.student.full_name).border = border
            ws.cell(row=row_num, column=3, value=f"{income.student.grade.number}{income.student.grade.parallel}").border = border
            ws.cell(row=row_num, column=4, value=float(income.amount)).border = border
            ws.cell(row=row_num, column=5, value=income.get_payment_method_display()).border = border
            ws.cell(row=row_num, column=6, value=income.get_status_display()).border = border
            ws.cell(row=row_num, column=7, value=income.transaction_id).border = border
            ws.cell(row=row_num, column=8, value=income.get_paid_months_display()).border = border

        # Добавляем строку с итогами
        last_row = len(queryset) + 2
        ws.cell(row=last_row, column=1, value="ИТОГО:").font = header_font
        ws.cell(row=last_row, column=4, value=float(total_amount)).font = header_font
        ws.cell(row=last_row, column=4).number_format = '#,##0.00'

        # Настраиваем ширину столбцов
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

        # Формируем ответ
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename="income_report_{datetime.now().strftime("%Y-%m-%d")}.xlsx"'},
        )
        wb.save(response)
        return response

   

    def get_queryset(self):
        queryset = super().get_queryset().select_related('student', 'academic_year')

        # Фильтрация по учебному году (по умолчанию — текущий)
        year_id = self.request.GET.get('year_id')
        if year_id == 'all':
            pass  # показываем все годы
        elif year_id:
            queryset = queryset.filter(academic_year_id=year_id)
        else:
            current_year = AcademicYear.objects.filter(is_current=True).first()
            if current_year:
                queryset = queryset.filter(academic_year=current_year)

        # Фильтрация по дате
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')

        if date_from:
            queryset = queryset.filter(date__gte=date_from)
        if date_to:
            queryset = queryset.filter(date__lte=date_to)

        # Фильтрация по статусу
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)

        # Фильтрация по способу оплаты
        payment_method = self.request.GET.get('payment_method')
        if payment_method:
            queryset = queryset.filter(payment_method=payment_method)

        # Фильтрация по ученику (через поиск)
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(student__full_name__icontains=search) |
                Q(transaction_id__icontains=search)
            )

        return queryset.order_by('-date')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Параметры фильтрации
        context['date_from'] = self.request.GET.get('date_from', '')
        context['date_to'] = self.request.GET.get('date_to', '')
        context['status'] = self.request.GET.get('status', '')
        context['payment_method'] = self.request.GET.get('payment_method', '')
        context['search'] = self.request.GET.get('search', '')

        # Учебный год
        year_id = self.request.GET.get('year_id', '')
        context['selected_year_id'] = year_id
        context['academic_years_list'] = AcademicYear.objects.order_by('-start_date')
        if not year_id:
            context['selected_year_id'] = str(
                (AcademicYear.objects.filter(is_current=True).first() or AcademicYear()).pk or ''
            )

        # Choices для фильтров
        context['status_choices'] = Income.STATUS_CHOICES
        context['payment_method_choices'] = Income.PAYMENT_METHODS

        # Суммарная информация
        queryset = self.get_queryset()
        context['total_amount'] = queryset.aggregate(Sum('amount'))['amount__sum'] or 0
        context['total_count'] = queryset.count()

        return context
    
  
    
    def export_to_csv(self):
        queryset = self.get_queryset()
        
        response = HttpResponse(
            content_type='text/csv',
            headers={'Content-Disposition': f'attachment; filename="income_report_{datetime.now().strftime("%Y-%m-%d")}.csv"'},
        )
        
        writer = csv.writer(response)
        
        # Заголовки CSV
        writer.writerow([
            'Дата',
            'Ученик',
            'Класс',
            'Сумма (сом)',
            'Способ оплаты',
            'Статус',
            'Номер транзакции',
            'Период оплаты',
            'Учебный год'
        ])
        
        # Данные
        for income in queryset:
            writer.writerow([
                income.date.strftime('%d.%m.%Y'),
                income.student.full_name,
                f"{income.student.grade.number}{income.student.grade.parallel}",
                income.amount,
                income.get_payment_method_display(),
                income.get_status_display(),
                income.transaction_id,
                income.get_paid_months_display(),
                income.academic_year.year if income.academic_year else ''
            ])
        
        return response



class DownloadReceiptView(View):
    def get(self, request, pk):
        from .models import Income
        income = Income.objects.get(pk=pk)
        pdf_buffer = generate_receipt_pdf(income)
        response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="receipt_{income.transaction_id}.pdf"'
        return response

      
class IncomeCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = Income
    form_class = IncomeForm
    template_name = 'school/incomes/form.html'
    
    def get_success_url(self):
        return reverse_lazy('student-detail', kwargs={'pk': self.kwargs['student_id']})
    def test_func(self):
        return True 
    
    def get_initial(self):
        initial = super().get_initial()
        student = get_object_or_404(Student, pk=self.kwargs['student_id'])
        initial.update({
            'student': student,
            'income_type': 'Оплата контракта',
            'date': timezone.now().date()
        })
        return initial
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['student_id'] = self.kwargs['student_id']
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        # Только суперпользователь выбирает дату, остальным ставится сегодня
        if not self.request.user.is_superuser:
            form.instance.date = timezone.now().date()
        # super() сохраняет форму один раз и устанавливает self.object
        response = super().form_valid(form)
        messages.success(
            self.request,
            f"Платеж успешно сохранен. Номер транзакции: {self.object.transaction_id}"
        )
        return response
    
    def get_success_url(self):
        # После скачивания перенаправляем на страницу ученика
        return reverse('student-detail', kwargs={'pk': self.kwargs['student_id']})
    def send_receipt_email(self, income):
        # Получаем email из контактов родителя (нужно адаптировать под вашу структуру)
        email = self.extract_email_from_contacts(income.student.parent_contacts)
        
        if email:
            try:
                subject = f"Квитанция об оплате #{income.transaction_id}"
                message = f"Уважаемые родители!\n\nПрикрепляем квитанцию об оплате для {income.student.full_name}."
                
                # Генерация HTML квитанции и конвертация в PDF
                # html_receipt = generate_html_receipt(income)
                # pdf_content = generate_pdf_from_html(html_receipt)
                
                email = EmailMessage(
                    subject,
                    message,
                    settings.DEFAULT_FROM_EMAIL,
                    [email]
                )
                email.attach(
                    f"receipt_{income.transaction_id}.pdf",
                    # pdf_content,
                    'application/pdf'
                )
                email.send()
            except Exception as e:
                print(f"Ошибка отправки email: {e}")
    
    def extract_email_from_contacts(self, contacts):
        # Простая реализация - нужно адаптировать под ваш формат хранения контактов
        import re
        email_match = re.search(r'[\w\.-]+@[\w\.-]+', contacts)
        return email_match.group(0) if email_match else None
    
    def download_receipt(self, pdf_buffer):
        response = HttpResponse(pdf_buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="receipt_{self.object.transaction_id}.pdf"'
        return response

from django.shortcuts import get_object_or_404
from num2words import num2words

class ReceiptPrintView(LoginRequiredMixin, DetailView):
    template_name = 'school/receipt_print.html'
    
    def get_object(self):
        student = get_object_or_404(Student, pk=self.kwargs['student_id'])
        payment = get_object_or_404(Income, pk=self.kwargs['payment_id'])
        return {'student': student, 'payment': payment}
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        student = self.object['student']
        payment = self.object['payment']
        context['student'] = student
        context['payment'] = payment

        # Сумма прописью
        som_int = int(payment.amount)
        tiyin = int(round((payment.amount - som_int) * 100))
        context['amount_words'] = f"{num2words(som_int, lang='ru').capitalize()} сом {tiyin:02d} тыйын"

        # Текущий учебный год
        current_year = AcademicYear.objects.filter(is_current=True).first()
        
        
        if current_year:
            
            payments = Income.objects.filter(
                student=student,
                academic_year=current_year,
                status='paid'
            ).aggregate(total=Sum('amount'))['total'] or 0
    
            
            
            
            contract_amount = student.contract_amount or 0
            remaining = contract_amount - payments
            context['remaining_payment'] = max(remaining, 0)
        else:
            context['remaining_payment'] = "Не установлен текущий учебный год"
        
        # Оплаченные месяцы
        if payment.paid_months:
            month_names = dict(Income.MONTH_CHOICES)
            paid_months = [month_names[int(m)] for m in payment.paid_months]
            context['paid_months'] = ", ".join(paid_months)
        else:
            context['paid_months'] = None
            
        return context





from django.db.models import Sum, Q
from datetime import datetime, timedelta
class ExpenseListView(LoginRequiredMixin, ListView):
    model = Expense
    template_name = 'school/expenses/list.html'
    context_object_name = 'expenses'
    paginate_by = 20
    
    def test_func(self):
        return is_admin(self.request.user) or is_accountant(self.request.user)
    
    def get_queryset(self):
        queryset = super().get_queryset().select_related('created_by', 'academic_year')

        # Фильтрация по учебному году (по умолчанию — текущий)
        year_id = self.request.GET.get('year_id')
        if year_id == 'all':
            pass
        elif year_id:
            queryset = queryset.filter(academic_year_id=year_id)
        else:
            current_year = AcademicYear.objects.filter(is_current=True).first()
            if current_year:
                queryset = queryset.filter(academic_year=current_year)

        # Фильтрация по дате
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        
        if date_from:
            queryset = queryset.filter(date__gte=date_from)
        if date_to:
            queryset = queryset.filter(date__lte=date_to)
            
        # Фильтрация по категории
        category = self.request.GET.get('category')
        if category:
            queryset = queryset.filter(category=category)
            
        # Фильтрация по способу оплаты
        payment_method = self.request.GET.get('payment_method')
        if payment_method:
            queryset = queryset.filter(payment_method=payment_method)
            
        # Фильтрация по поставщику (поиск)
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(supplier__icontains=search) |
                Q(notes__icontains=search) |
                Q(invoice_number__icontains=search)
            )
            
        # Сортировка
        sort = self.request.GET.get('sort', '-date')
        queryset = queryset.order_by(sort)
        
        return queryset
    
    def render_to_response(self, context, **response_kwargs):
        if self.request.GET.get('export') == 'excel':
            return self.export_to_excel()
        return super().render_to_response(context, **response_kwargs)
    def export_to_excel(self):
        queryset = self.get_queryset()
        date_from = self.request.GET.get('date_from', 'не указана')
        date_to = self.request.GET.get('date_to', 'не указана')
    
    # Создаем Excel-файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Отчет по расходам"
    
    # Стили
        header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        header_font = Font(bold=True)
        center_aligned = Alignment(horizontal='center')
        right_aligned = Alignment(horizontal='right')
        border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
        )

    # Заголовок отчета
        ws.append(["Отчет по расходам"])
        ws.merge_cells('A1:G1')
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = center_aligned
    
        ws.append([f"Период: с {date_from} по {date_to}"])
        ws.merge_cells('A2:G2')
        ws['A2'].alignment = center_aligned
    
        ws.append([])  # Пустая строка

    # Заголовки таблицы
        headers = [
        'Дата', 'Категория', 'Поставщик', 
        'Сумма (сом)', 'Способ оплаты', 
        'Номер счета', 'Примечания'
    ]
    
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_aligned
            cell.border = border

    # Данные
        for row_num, expense in enumerate(queryset, 5):
            ws.cell(row=row_num, column=1, value=expense.date).border = border
            ws.cell(row=row_num, column=2, value=expense.get_category_display()).border = border
            ws.cell(row=row_num, column=3, value=expense.supplier).border = border
            ws.cell(row=row_num, column=4, value=float(expense.amount)).border = border
            ws.cell(row=row_num, column=5, value=expense.get_payment_method_display()).border = border
            ws.cell(row=row_num, column=6, value=expense.invoice_number or '').border = border
            ws.cell(row=row_num, column=7, value=expense.notes or '').border = border

        # Форматируем числовые ячейки
            ws.cell(row=row_num, column=4).number_format = '#,##0.00'
            ws.cell(row=row_num, column=4).alignment = right_aligned

    # Итоговая строка
        last_row = len(queryset) + 5
        ws.cell(row=last_row, column=3, value="ИТОГО:").font = header_font
        ws.cell(row=last_row, column=4, 
           value=f"=SUM(D5:D{last_row-1})").font = header_font
        ws.cell(row=last_row, column=4).number_format = '#,##0.00'
        ws.cell(row=last_row, column=4).alignment = right_aligned

    # Настраиваем ширину столбцов (только для столбцов с данными, пропуская объединенные ячейки)
        for col in ws.iter_cols(min_row=4, max_row=ws.max_row, min_col=1, max_col=7):
            max_length = 0
            column_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width

    # Формируем ответ
        response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename="expense_report.xlsx"'},
    )
        wb.save(response)
        return response
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Параметры фильтрации
        context['date_from'] = self.request.GET.get('date_from', '')
        context['date_to'] = self.request.GET.get('date_to', '')
        context['category'] = self.request.GET.get('category', '')
        context['payment_method'] = self.request.GET.get('payment_method', '')
        context['search'] = self.request.GET.get('search', '')
        context['sort'] = self.request.GET.get('sort', '-date')

        # Учебный год
        year_id = self.request.GET.get('year_id', '')
        context['selected_year_id'] = year_id
        context['academic_years_list'] = AcademicYear.objects.order_by('-start_date')
        if not year_id:
            current = AcademicYear.objects.filter(is_current=True).first()
            context['selected_year_id'] = str(current.pk) if current else ''

        # Варианты для фильтров
        context['category_choices'] = Expense.CATEGORIES
        context['payment_method_choices'] = Expense.PAYMENT_METHODS

        # Статистика
        queryset = self.get_queryset()
        context['total_amount'] = queryset.aggregate(Sum('amount'))['amount__sum'] or 0
        context['expenses_count'] = queryset.count()

        # Даты по умолчанию (последние 30 дней)
        today = datetime.datetime.now().date()
        context['default_date_from'] = (today - timedelta(days=30)).strftime('%Y-%m-%d')
        context['default_date_to'] = today.strftime('%Y-%m-%d')

        return context
 
    
class ExpenseCreateView(LoginRequiredMixin, CreateView):
    model = Expense
    form_class = ExpenseForm
    template_name = 'school/expenses/form.html'
    success_url = reverse_lazy('expense-list')
    
    def test_func(self):
        return is_admin(self.request.user) or is_accountant(self.request.user)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Добавление расхода'
        context['submit_text'] = 'Добавить расход'
        return context
    
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(
            self.request,
            f"Расход на сумму {form.instance.amount} сом успешно добавлен!"
        )
        return super().form_valid(form)


# Отчеты
@login_required
@user_passes_test(lambda u: is_admin(u) or is_accountant(u))
def reports(request):
    return render(request, 'school/reports/index.html')

# сотрудники
class EmployeeCreateView(LoginRequiredMixin, CreateView):
    model = Employee
    form_class = EmployeeForm
    template_name = 'school/employees/form.html'
    success_url = reverse_lazy('employee-list')
    
    def test_func(self):
        return is_admin(self.request.user)
    
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, "Сотрудник успешно добавлен")
        return super().form_valid(form)  # Это вернет HttpResponseRedirect
    
    def form_invalid(self, form):
        return self.render_to_response(self.get_context_data(form=form))  # Вернет HttpResponse


class EmployeeUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Employee
    form_class = EmployeeForm
    template_name = 'school/employees/form.html'
    success_url = reverse_lazy('employee-list')
    
    def test_func(self):
        return is_admin(self.request.user)
    
    def form_valid(self, form):
        messages.success(self.request, "Данные сотрудника обновлены")
        return super().form_valid(form)

class EmployeeListView(LoginRequiredMixin, ListView):
    model = Employee
    template_name = 'school/employees/list.html'
    context_object_name = 'employees'
    
    def test_func(self):
        return is_admin(self.request.user) or is_accountant(self.request.user)
    
    def get_queryset(self):
        return Employee.objects.filter(is_active=True).select_related('position')
    
    
class SalaryPaymentCreateView(LoginRequiredMixin, CreateView):
    model = SalaryPayment
    form_class = SalaryPaymentForm
    template_name = 'school/employees/salary_payment_form.html'
    
    def test_func(self):
        return is_admin(self.request.user) or is_accountant(self.request.user)
    
    def get_initial(self):
        initial = super().get_initial()
        employee_id = self.request.GET.get('employee')
        if employee_id:
            try:
                employee = Employee.objects.get(pk=employee_id)
                initial['employee'] = employee
                initial['amount'] = employee.monthly_salary
            except Employee.DoesNotExist:
                pass
        return initial
    
    def get_success_url(self):
        employee_id = self.request.GET.get('employee')
        if employee_id:
            return reverse('employee-detail', kwargs={'pk': employee_id})
        return reverse('employee-list')
    
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(
            self.request,
            f"Зарплатная выплата для {form.instance.employee} успешно добавлена!"
        )
        return super().form_valid(form)

class SalaryReportView(LoginRequiredMixin, ListView):
    template_name = 'school/employees/salary_report.html'
    context_object_name = 'payments'
    
    def test_func(self):
        return is_admin(self.request.user) or is_accountant(self.request.user)
    
    def get_queryset(self):
        queryset = SalaryPayment.objects.select_related('employee', 'employee__position', 'created_by')
        
        year = self.request.GET.get('year')
        month = self.request.GET.get('month')
        
        if year:
            queryset = queryset.filter(for_month__year=year)
        if month:
            queryset = queryset.filter(for_month__month=month)
            
        return queryset.order_by('-for_month', 'employee__full_name')
    
    def render_to_response(self, context, **response_kwargs):
        if self.request.GET.get('export') == 'xlsx':
            return self.export_to_excel()
        return super().render_to_response(context, **response_kwargs)

    def export_to_excel(self):
        queryset = self.get_queryset()
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=salary_report.xlsx'
    
        wb = Workbook()
        ws = wb.active
        ws.title = "Зарплатные выплаты"
    
    # Заголовки
        headers = ['Сотрудник', 'Должность', 'Месяц', 'Сумма', 'Дата выплаты', 'Способ оплаты', 'Тип']
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header).font = Font(bold=True)
    
    # Данные
        for row_num, payment in enumerate(queryset, 2):
            ws.cell(row=row_num, column=1, value=payment.employee.full_name)
            ws.cell(row=row_num, column=2, value=str(payment.employee.position))
            ws.cell(row=row_num, column=3, value=payment.for_month.strftime('%B %Y'))
            ws.cell(row=row_num, column=4, value=float(payment.amount))
            ws.cell(row=row_num, column=5, value=payment.payment_date.strftime('%d.%m.%Y'))
            ws.cell(row=row_num, column=6, value=payment.get_payment_method_display())
            ws.cell(row=row_num, column=7, value='Премия' if payment.is_bonus else 'Зарплата')
    
    # Автоширина столбцов
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width
    
        wb.save(response)
        return response
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Добавляем доступные годы для фильтра
        years = SalaryPayment.objects.dates('for_month', 'year').order_by('-for_month')
        unique_years = sorted(set([year.year for year in years]), reverse=True)
        context['years'] = unique_years
        context['selected_year'] = self.request.GET.get('year')

        
        # Добавляем месяцы для фильтра
        context['months'] = [
            (1, 'Январь'), (2, 'Февраль'), (3, 'Март'), 
            (4, 'Апрель'), (5, 'Май'), (6, 'Июнь'),
            (7, 'Июль'), (8, 'Август'), (9, 'Сентябрь'),
            (10, 'Октябрь'), (11, 'Ноябрь'), (12, 'Декабрь')
        ]
        context['selected_month'] = self.request.GET.get('month')
        
        # Суммарная информация
        queryset = self.get_queryset()
        context['total_amount'] = queryset.aggregate(
            Sum('amount')
        )['amount__sum'] or 0
        
        return context      
    
    
class EmployeeDetailView(LoginRequiredMixin, DetailView):
    model = Employee
    template_name = 'school/employees/detail.html'
    
    def test_func(self):
        return is_admin(self.request.user) or is_accountant(self.request.user)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['salary_payments'] = SalaryPayment.objects.filter(
            employee=self.object
        ).order_by('-for_month')
        return context


from django.views.generic import ListView
from .models import Document

class DocumentListView(ListView):
    model = Document
    template_name = 'documents.html'
    context_object_name = 'documents'
    
    def get_queryset(self):
        # Группируем документы по категориям
        documents = super().get_queryset()
        categories = {}
        for doc in documents:
            if doc.category not in categories:
                categories[doc.category] = []
            categories[doc.category].append(doc)
        return categories
    
from django.views.generic import ListView, DetailView
from .models import GalleryEvent
from django.views.generic import ListView
from .models import GalleryEvent
from django.db.models import Count
from django.views.generic import ListView
from .models import GalleryEvent
from django.db.models import Count

class GalleryListView(ListView):
    model = GalleryEvent
    template_name = 'gallery.html'
    context_object_name = 'events'
    paginate_by = 12
    
    def get_queryset(self):
        # Получаем базовый queryset
        queryset = GalleryEvent.objects.annotate(
            image_count=Count('images')
        ).filter(image_count__gt=0)
        
        # Фильтрация по году
        year = self.request.GET.get('year')
        if year and year != 'all':
            queryset = queryset.filter(date__year=year)
        
        return queryset.order_by('-date')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Получаем уникальные года из событий
        years = GalleryEvent.objects.annotate(
            image_count=Count('images')
        ).filter(image_count__gt=0).values_list('date__year', flat=True).distinct().order_by('-date__year')
        
        # Добавляем в контекст
        context['years'] = years
        context['current_year'] = self.request.GET.get('year', 'all')
        
        return context
class GalleryListView(ListView):
    model = GalleryEvent
    template_name = 'gallery.html'
    context_object_name = 'events'
    paginate_by = 12
    
    def get_queryset(self):
        # Убираем аннотацию image_count
        queryset = GalleryEvent.objects.filter(images__isnull=False).distinct()
        
        # Фильтрация по году
        year = self.request.GET.get('year')
        if year and year != 'all':
            queryset = queryset.filter(date__year=year)
        
        return queryset.order_by('-date')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Получаем уникальные года из событий
        years = GalleryEvent.objects.values_list('date__year', flat=True).distinct().order_by('-date__year')
        
        # Добавляем в контекст
        context['years'] = years
        context['current_year'] = self.request.GET.get('year', 'all')
        
        return context
    
    

def teacher_view1(request):
    model = Teacher.objects.all()
    return render(request,'teacher.html',{'teachers':model})
class GalleryDetailView(DetailView):
    model = GalleryEvent
    template_name = 'gallery_detail.html'
    context_object_name = 'event'
    slug_field = 'slug'
    slug_url_kwarg = 'slug'
    
from django.views.generic import DetailView
from django.template.loader import get_template
from xhtml2pdf import pisa
from django.http import HttpResponse
import os
from django.template.loader import get_template
from django.http import HttpResponse
from xhtml2pdf import pisa
from io import BytesIO
# import datetime
from django.template.loader import get_template
from django.http import HttpResponse
from xhtml2pdf import pisa
from io import BytesIO
import datetime
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont




from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from .pdf_utils import register_fonts
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet

from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm

from num2words import num2words
def expense_receipt_pdf(request, pk):
    from django.http import HttpResponse
    from django.shortcuts import get_object_or_404
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from num2words import num2words

    expense = get_object_or_404(Expense, pk=pk)

    # твоя функция регистрации шрифтов
    register_fonts()

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="receipt_{expense.id}.pdf"'

    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )

    elements = []
    styles = getSampleStyleSheet()

    style_heading = ParagraphStyle(
        "Heading1",
        parent=styles["Heading1"],
        fontName="Arial",
        fontSize=10,
        alignment=1,  # center
        spaceAfter=6,
    )

    style_normal = ParagraphStyle(
        "Normal",
        parent=styles["Normal"],
        fontName="Arial",
        fontSize=10,
        leading=12,
    )

    # --- Шапка
    elements.append(Paragraph("РАСХОДНЫЙ КАССОВЫЙ ОРДЕР", style_heading))
    elements.append(
        Paragraph(
            f"№ {expense.id or 'БН'} от {expense.created_at.strftime('%d.%m.%Y')}",
            style_normal,
        )
    )
    elements.append(Spacer(1, 0.2 * cm))

    # --- сумма с разделением по 3 символа
    formatted_amount = "{:,.2f}".format(float(expense.amount)).replace(",", " ")

    # --- сумма прописью
    som_int = int(float(expense.amount))
    tiyin = int(round((float(expense.amount) - som_int) * 100))
    amount_words = f"{num2words(som_int, lang='ru').capitalize()} сом {tiyin:02d} тыйын"

    # --- кто оформил
    issuer = "Не указан"
    if getattr(request, "user", None) and getattr(request.user, "is_authenticated", False):
        issuer = (
            getattr(request.user, "get_full_name", lambda: "")()  # type: ignore
            or getattr(request.user, "username", "")
            or str(request.user)
        )

    # --- получатель (попробуем угадать по разным полям, чтобы не падало)
    recipient_fio = ""
    for attr in ("recipient",   "receiver", "employee", "worker", "person", "user"):
        obj = getattr(expense, attr, None)
        if obj:
            recipient_fio = (
                getattr(obj, "get_full_name", lambda: "")()
                or getattr(obj, "full_name", "")
                or getattr(obj, "fio", "")
                or getattr(obj, "name", "")
                or str(obj)
            )
            break
    # если есть строковое поле (например receiver_name / recipient_name)
    for attr in ("recipient_name", "receiver_name", "fio", "full_name"):
        val = getattr(expense, attr, None)
        if val and not recipient_fio:
            recipient_fio = str(val).strip()
            break
    if not recipient_fio:
        recipient_fio = "ФИО"

    # --- Таблица реквизитов
    data = [
        ["Дата расхода:", expense.date.strftime("%d.%m.%Y") if expense.date else ""],
        ["Номер документа:", str(expense.id)],
        ["Категория расхода:", expense.get_category_display()],
        ["Поставщик:", expense.supplier],
        ["Сумма расхода:", f"{formatted_amount} сом"],
        ["Сумма прописью:", amount_words],
        ["Способ оплаты:", expense.get_payment_method_display()],
        ["Основание:", expense.notes or "Оплата услуг"],
        ["Оформил:", issuer],
    ]

    table = Table(data, colWidths=[5 * cm, 12 * cm])
    table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), "Arial"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (0, 0), (0, -1), "RIGHT"),
                ("ALIGN", (1, 0), (1, -1), "LEFT"),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("GRID", (0, 0), (-1, -1), 0.7, colors.black),
            ]
        )
    )

    elements.append(table)
    elements.append(Spacer(1, 0.5 * cm))

    # --- Красивые подписи (линия под подпись реальная, а не ____)

    def sign_cell(title: str, fio: str):
        t = Table(
            [
                [Paragraph(f"{title}:", style_normal)],
                [""],  # линия
                [Paragraph(fio or "ФИО", style_normal)],
            ],
            colWidths=[8.5 * cm],
            rowHeights=[0.55 * cm, 0.9 * cm, 0.55 * cm],
        )
        t.setStyle(
            TableStyle(
                [
                    ("FONTNAME", (0, 0), (-1, -1), "Arial"),
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),

                    ("ALIGN", (0, 0), (-1, 0), "LEFT"),
                    ("LINEBELOW", (0, 1), (0, 1), 1, colors.black),  # линия подписи
                    ("ALIGN", (0, 2), (-1, 2), "CENTER"),

                    ("TOPPADDING", (0, 0), (-1, -1), 0),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                    ("LEFTPADDING", (0, 0), (-1, -1), 0),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ]
            )
        )
        return t

    director_fio = "Муртазо У.Б."

    signatures = Table(
        [[sign_cell("Получатель", recipient_fio), sign_cell("Директор", director_fio)]],
        colWidths=[9 * cm, 9 * cm],
    )
    signatures.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )

    elements.append(signatures)

    doc.build(elements)
    return response




from .models import Graduate

from django.shortcuts import render
from .models import Graduate
from django.views.generic import ListView
from .models import Application
def graduates_list(request):
    graduates = Graduate.objects.all().order_by('-graduation_year', '-order', 'username')
    unique_years = Graduate.objects.values_list('graduation_year', flat=True).distinct().order_by('-graduation_year')
    
    # Фильтрация по году, если указан параметр
    year_filter = request.GET.get('year')
    if year_filter:
        graduates = graduates.filter(graduation_year=year_filter)
    
    return render(request, 'school/graduates/graduates_list.html', {
        'graduates': graduates,
        'unique_years': unique_years,
        'selected_year': year_filter,
    })


# ══════════════════════════════════════════════════════════════════
#  CMS — управление контентом сайта
# ══════════════════════════════════════════════════════════════════
from .models import News, Teacher, Student2, Graduate, GalleryEvent
from .forms import (NewsForm, TeacherForm, BestStudentForm,
                    GraduateForm, GalleryEventForm, GalleryImageFormSet)


def _has_cms_access(user):
    """Доступ к CMS: staff, superuser или СММ-пользователь."""
    if user.is_staff or user.is_superuser:
        return True
    return getattr(getattr(user, 'profile', None), 'is_cms_user', False)


def cms_login(request):
    if request.user.is_authenticated and _has_cms_access(request.user):
        return redirect('cms-dashboard')
    error = None
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        user = authenticate(request, username=username, password=password)
        if user is not None and _has_cms_access(user):
            login(request, user)
            return redirect(request.GET.get('next', 'cms-dashboard'))
        else:
            error = 'Неверный логин/пароль или нет доступа к CMS.'
    return render(request, 'school/cms/login.html', {'error': error})


def cms_logout(request):
    logout(request)
    return redirect('cms-login')


class CMSRequiredMixin(LoginRequiredMixin):
    login_url = '/cms/login/'

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect(f'/cms/login/?next={request.path}')
        if not _has_cms_access(request.user):
            return redirect(f'/cms/login/?next={request.path}')
        return super(LoginRequiredMixin, self).dispatch(request, *args, **kwargs)


@login_required(login_url='/cms/login/')
def cms_dashboard(request):
    if not _has_cms_access(request.user):
        return redirect('cms-login')
    context = {
        'news_count':         News.objects.count(),
        'teacher_count':      Teacher.objects.count(),
        'best_student_count': Student2.objects.count(),
        'graduate_count':     Graduate.objects.count(),
        'gallery_count':      GalleryEvent.objects.count(),
        'recent_news':        News.objects.order_by('-created_at')[:3],
    }
    return render(request, 'school/cms/dashboard.html', context)


# ── Новости ──────────────────────────────────────────────────────

class CMSNewsListView(CMSRequiredMixin, ListView):
    model = News
    template_name = 'school/cms/news/list.html'
    context_object_name = 'items'
    ordering = ['-created_at']
    paginate_by = 20


class CMSNewsCreateView(CMSRequiredMixin, CreateView):
    model = News
    form_class = NewsForm
    template_name = 'school/cms/form.html'
    success_url = reverse_lazy('cms-news-list')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['title'] = 'Добавить новость'
        ctx['back_url'] = reverse_lazy('cms-news-list')
        return ctx


class CMSNewsUpdateView(CMSRequiredMixin, UpdateView):
    model = News
    form_class = NewsForm
    template_name = 'school/cms/form.html'
    success_url = reverse_lazy('cms-news-list')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['title'] = 'Редактировать новость'
        ctx['back_url'] = reverse_lazy('cms-news-list')
        return ctx


class CMSNewsDeleteView(CMSRequiredMixin, DeleteView):
    model = News
    template_name = 'school/cms/confirm_delete.html'
    success_url = reverse_lazy('cms-news-list')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['title'] = 'Удалить новость'
        ctx['object_name'] = str(self.object)
        ctx['back_url'] = reverse_lazy('cms-news-list')
        return ctx


# ── Учителя ──────────────────────────────────────────────────────

class CMSTeacherListView(CMSRequiredMixin, ListView):
    model = Teacher
    template_name = 'school/cms/teachers/list.html'
    context_object_name = 'items'
    ordering = ['order', 'last_name']


class CMSTeacherCreateView(CMSRequiredMixin, CreateView):
    model = Teacher
    form_class = TeacherForm
    template_name = 'school/cms/form.html'
    success_url = reverse_lazy('cms-teacher-list')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['title'] = 'Добавить учителя'
        ctx['back_url'] = reverse_lazy('cms-teacher-list')
        return ctx


class CMSTeacherUpdateView(CMSRequiredMixin, UpdateView):
    model = Teacher
    form_class = TeacherForm
    template_name = 'school/cms/form.html'
    success_url = reverse_lazy('cms-teacher-list')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['title'] = 'Редактировать учителя'
        ctx['back_url'] = reverse_lazy('cms-teacher-list')
        return ctx


class CMSTeacherDeleteView(CMSRequiredMixin, DeleteView):
    model = Teacher
    template_name = 'school/cms/confirm_delete.html'
    success_url = reverse_lazy('cms-teacher-list')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['title'] = 'Удалить учителя'
        ctx['object_name'] = str(self.object)
        ctx['back_url'] = reverse_lazy('cms-teacher-list')
        return ctx


# ── Лучшие ученики ───────────────────────────────────────────────

class CMSBestStudentListView(CMSRequiredMixin, ListView):
    model = Student2
    template_name = 'school/cms/best_students/list.html'
    context_object_name = 'items'
    ordering = ['order', 'last_name']


class CMSBestStudentCreateView(CMSRequiredMixin, CreateView):
    model = Student2
    form_class = BestStudentForm
    template_name = 'school/cms/form.html'
    success_url = reverse_lazy('cms-best-student-list')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['title'] = 'Добавить лучшего ученика'
        ctx['back_url'] = reverse_lazy('cms-best-student-list')
        return ctx


class CMSBestStudentUpdateView(CMSRequiredMixin, UpdateView):
    model = Student2
    form_class = BestStudentForm
    template_name = 'school/cms/form.html'
    success_url = reverse_lazy('cms-best-student-list')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['title'] = 'Редактировать лучшего ученика'
        ctx['back_url'] = reverse_lazy('cms-best-student-list')
        return ctx


class CMSBestStudentDeleteView(CMSRequiredMixin, DeleteView):
    model = Student2
    template_name = 'school/cms/confirm_delete.html'
    success_url = reverse_lazy('cms-best-student-list')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['title'] = 'Удалить лучшего ученика'
        ctx['object_name'] = str(self.object)
        ctx['back_url'] = reverse_lazy('cms-best-student-list')
        return ctx


# ── Выпускники ───────────────────────────────────────────────────

class CMSGraduateListView(CMSRequiredMixin, ListView):
    model = Graduate
    template_name = 'school/cms/graduates/list.html'
    context_object_name = 'items'
    ordering = ['-graduation_year', '-order']
    paginate_by = 20


class CMSGraduateCreateView(CMSRequiredMixin, CreateView):
    model = Graduate
    form_class = GraduateForm
    template_name = 'school/cms/form.html'
    success_url = reverse_lazy('cms-graduate-list')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['title'] = 'Добавить выпускника'
        ctx['back_url'] = reverse_lazy('cms-graduate-list')
        return ctx


class CMSGraduateUpdateView(CMSRequiredMixin, UpdateView):
    model = Graduate
    form_class = GraduateForm
    template_name = 'school/cms/form.html'
    success_url = reverse_lazy('cms-graduate-list')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['title'] = 'Редактировать выпускника'
        ctx['back_url'] = reverse_lazy('cms-graduate-list')
        return ctx


class CMSGraduateDeleteView(CMSRequiredMixin, DeleteView):
    model = Graduate
    template_name = 'school/cms/confirm_delete.html'
    success_url = reverse_lazy('cms-graduate-list')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['title'] = 'Удалить выпускника'
        ctx['object_name'] = str(self.object)
        ctx['back_url'] = reverse_lazy('cms-graduate-list')
        return ctx


# ── Галерея ──────────────────────────────────────────────────────

class CMSGalleryListView(CMSRequiredMixin, ListView):
    model = GalleryEvent
    template_name = 'school/cms/gallery/list.html'
    context_object_name = 'items'
    ordering = ['-date']
    paginate_by = 20


class CMSGalleryCreateView(CMSRequiredMixin, CreateView):
    model = GalleryEvent
    form_class = GalleryEventForm
    template_name = 'school/cms/gallery/form.html'
    success_url = reverse_lazy('cms-gallery-list')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['title'] = 'Добавить событие галереи'
        ctx['formset'] = GalleryImageFormSet(self.request.POST or None,
                                             self.request.FILES or None)
        return ctx

    def form_valid(self, form):
        ctx = self.get_context_data()
        formset = ctx['formset']
        if formset.is_valid():
            self.object = form.save()
            formset.instance = self.object
            formset.save()
            return redirect(self.success_url)
        return self.render_to_response(ctx)


class CMSGalleryUpdateView(CMSRequiredMixin, UpdateView):
    model = GalleryEvent
    form_class = GalleryEventForm
    template_name = 'school/cms/gallery/form.html'
    success_url = reverse_lazy('cms-gallery-list')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['title'] = 'Редактировать событие галереи'
        ctx['formset'] = GalleryImageFormSet(
            self.request.POST or None,
            self.request.FILES or None,
            instance=self.object,
        )
        return ctx

    def form_valid(self, form):
        ctx = self.get_context_data()
        formset = ctx['formset']
        if formset.is_valid():
            self.object = form.save()
            formset.instance = self.object
            formset.save()
            return redirect(self.success_url)
        return self.render_to_response(ctx)


class CMSGalleryDeleteView(CMSRequiredMixin, DeleteView):
    model = GalleryEvent
    template_name = 'school/cms/confirm_delete.html'
    success_url = reverse_lazy('cms-gallery-list')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['title'] = 'Удалить событие галереи'
        ctx['object_name'] = str(self.object)
        ctx['back_url'] = reverse_lazy('cms-gallery-list')
        return ctx



class ApplicationListView(ListView):
    model = Application
    template_name = 'school/application/application_list.html'
    context_object_name = 'applications'
    paginate_by = 10  # Пагинация по 10 элементов
    
    def get_queryset(self):
        # Сортировка по дате создания (новые сначала)
        return Application.objects.all().order_by('-created_at')